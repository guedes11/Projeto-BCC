from extrai_tabelas import extrai_tabelas_da_url
from openpyxl import load_workbook
from os import listdir
import re
    
                
urls = [
        f"https://pt.wikipedia.org/wiki/Campeonato_Brasileiro_de_Futebol_de_{ano}_-_Série_A" 
        for ano in range(2013, 2024)
       ]

extrai_tabelas_da_url(urls)

caminho = './Tabelas'

# Percorrendo cada arquivo .xlsx da pasta Tabelas.
for ano, arquivo in enumerate(listdir(caminho)):

    # Abrindo o arquivo .xlsx atual.
    pasta_trabalho = load_workbook(caminho + '/' + arquivo, read_only=False, data_only=False)
    planilha = pasta_trabalho.active

    # Percorrendo cada linha e coluna do arquivo.
    for linha in range(2, planilha.max_row + 1):
        for coluna in range(1, planilha.max_column + 1):
            celula = planilha.cell(row=linha, column=coluna).value

            # Convertendo o Placar em pontuação.
            if celula != "—" and coluna > 2:
                gols_mandante, gols_visitante = placar = (3, 0) if celula == "W.O.[a]" else map(int, re.findall(r'\d+', celula))
                resultado = gols_mandante - gols_visitante

                planilha.cell(row=linha, column=coluna).value = 3 if resultado > 0 else 1 if resultado == 0 else 0

    # Salvando o arquivo .xlsx com os valores convertidos.
    pasta_trabalho.close()
    pasta_trabalho.save(caminho + "/" + arquivo)
